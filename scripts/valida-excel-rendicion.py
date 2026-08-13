"""Valida la planilla de rendición sin LibreOffice.

No hay soffice en esta máquina, así que en vez de recalcular con Excel se hace
lo equivalente a mano: se leen los valores literales de las celdas, se evalúan
las fórmulas del libro contra esos valores, y se compara con lo esperado.
"""
import re
import sys
import zipfile

import openpyxl

ruta = sys.argv[1]
wb = openpyxl.load_workbook(ruta)  # con fórmulas, sin data_only

fallos = []
def check(ok, msg):
    print(("  OK   " if ok else "  FALLA") + " " + msg)
    if not ok:
        fallos.append(msg)

print("HOJAS")
check(wb.sheetnames == ["Rendición", "Respaldos"], f"nombres exactos: {wb.sheetnames}")

r = wb["Rendición"]
sp = wb["Respaldos"]

print("\nENCABEZADO")
check(r["A1"].value == "RENDICIÓN DE FONDO POR RENDIR", "A1 título")
check(r["A2"].value == "Performance Technologies SpA", "A2 empresa")
check(r["D4"].value == "Alex Oliva", f"D4 empleado = {r['D4'].value!r}")
check(r["D8"].value == 100000, f"D8 fondo = {r['D8'].value!r}")
check(r["D7"].value == "2026-08-01 al 2026-08-02", f"D7 período = {r['D7'].value!r}")
check(r["D9"].value == "3 comprobantes", f"D9 = {r['D9'].value!r}")

print("\nCABECERAS DE TABLA (fila 11)")
esperadas = ["N°","Fecha","Proveedor","RUT","N° Documento","Tipo Documento",
             "Detalle del Gasto","Categoría","Neto","IVA","Total (CLP)","Respaldo"]
reales = [r.cell(row=11, column=c).value for c in range(1, 13)]
check(reales == esperadas, f"12 columnas A11:L11 = {reales}")

# --- Datos y evaluación de fórmulas -----------------------------------------
PRIMERA, N = 12, 3
ULTIMA = PRIMERA + N - 1
TOTAL_ROW = ULTIMA + 1

netos = [r.cell(row=f, column=9).value for f in range(PRIMERA, ULTIMA + 1)]
ivas = [r.cell(row=f, column=10).value for f in range(PRIMERA, ULTIMA + 1)]
totales = [r.cell(row=f, column=11).value for f in range(PRIMERA, ULTIMA + 1)]
cats = [r.cell(row=f, column=8).value for f in range(PRIMERA, ULTIMA + 1)]

print("\nFILAS DE DATOS")
check(netos == [42017, 18000, 4116], f"netos I12:I14 = {netos}")
check(ivas == [7983, 0, 782], f"IVA J12:J14 = {ivas}")
check(totales == [50000, 18000, 4898], f"totales K12:K14 = {totales}")
check(r["B14"].value == "[ilegible]", f"B14 fecha ilegible = {r['B14'].value!r}")
check(r["F12"].value == "Boleta Electrónica", f"F12 etiqueta tipo = {r['F12'].value!r}")

print("\nFILA DE TOTALES (fila %d) — fórmulas evaluadas a mano" % TOTAL_ROW)
check(r[f"A{TOTAL_ROW}"].value == "TOTALES", "etiqueta TOTALES")
for col, valores in (("I", netos), ("J", ivas), ("K", totales)):
    f = r[f"{col}{TOTAL_ROW}"].value
    esperado = f"=SUM({col}{PRIMERA}:{col}{ULTIMA})"
    check(f == esperado, f"{col}{TOTAL_ROW} = {f!r}  →  suma {sum(valores)}")

check(sum(netos) + sum(ivas) == sum(totales),
      f"SUM(I)+SUM(J) == SUM(K): {sum(netos)}+{sum(ivas)}=={sum(totales)}")

# --- Cuadro financiero ------------------------------------------------------
FIN1, FIN2, FIN3 = TOTAL_ROW + 2, TOTAL_ROW + 3, TOTAL_ROW + 4
print(f"\nCUADRO FINANCIERO (filas {FIN1}-{FIN3})")
check(r[f"K{FIN1}"].value == "=D8", f"K{FIN1} = {r[f'K{FIN1}'].value!r}")
check(r[f"K{FIN2}"].value == f"=K{TOTAL_ROW}", f"K{FIN2} = {r[f'K{FIN2}'].value!r}")
check(r[f"K{FIN3}"].value == f"=K{FIN2}-K{FIN1}", f"K{FIN3} = {r[f'K{FIN3}'].value!r}")
saldo = sum(totales) - 100000
check(saldo == -27102, f"saldo evaluado = {saldo}")
check(r[f"G{FIN3}"].value == "Saldo a reintegrar a la empresa:",
      f"etiqueta de saldo negativo = {r[f'G{FIN3}'].value!r}")

# --- Resumen por categoría --------------------------------------------------
CAT_TIT = FIN3 + 5
CAT_PRIM = CAT_TIT + 2
CATS = ["Combustible","Peajes","Alimentación","Alojamiento","Transporte","Insumos / Otros"]
print(f"\nRESUMEN POR CATEGORÍA (título fila {CAT_TIT})")
check(r[f"A{CAT_TIT}"].value == "RESUMEN POR CATEGORÍA", f"título = {r[f'A{CAT_TIT}'].value!r}")
nombres = [r.cell(row=CAT_PRIM + i, column=1).value for i in range(6)]
check(nombres == CATS, f"6 categorías = {nombres}")

suma_cat = 0
for i, c in enumerate(CATS):
    fila = CAT_PRIM + i
    formula = r.cell(row=fila, column=4).value
    esperada = f'=SUMIF($H${PRIMERA}:$H${ULTIMA},"{c}",$K${PRIMERA}:$K${ULTIMA})'
    # Evaluar el SUMIF a mano contra los valores reales de H y K
    valor = sum(t for cat, t in zip(cats, totales) if cat == c)
    suma_cat += valor
    check(formula == esperada, f"{c}: fórmula ok → {valor}")

CAT_ULT = CAT_PRIM + 5
CAT_TOT = CAT_ULT + 1
check(r[f"D{CAT_TOT}"].value == f"=SUM(D{CAT_PRIM}:D{CAT_ULT})", f"TOTAL categorías = {r[f'D{CAT_TOT}'].value!r}")
check(suma_cat == sum(totales),
      f"TOTAL categorías ({suma_cat}) cuadra con K{TOTAL_ROW} ({sum(totales)})")

# --- Resumen tributario -----------------------------------------------------
TRIB_TIT = CAT_TOT + 2
print(f"\nRESUMEN TRIBUTARIO (título fila {TRIB_TIT})")
check(r[f"A{TRIB_TIT}"].value == "RESUMEN TRIBUTARIO", f"título = {r[f'A{TRIB_TIT}'].value!r}")
exento_eval = sum(t for iva, t in zip(ivas, totales) if iva == 0)
check(r.cell(row=TRIB_TIT + 4, column=4).value ==
      f'=SUMIF($J${PRIMERA}:$J${ULTIMA},0,$K${PRIMERA}:$K${ULTIMA})',
      f"Total exento: fórmula ok → {exento_eval} (el pasaje terrestre)")
check(exento_eval == 18000, f"exento evaluado = {exento_eval}")
check(r.cell(row=TRIB_TIT + 5, column=1).value == "TOTAL RENDICIÓN", "última fila tributaria")

# Y lo que motivó todo esto: el neto y el IVA de las filas tienen que venir
# CALCULADOS, no en 0. Con los datos de prueba (una boleta con desglose, un
# exento y uno con campos ilegibles) ninguna fila afecta puede quedar en 0.
netos_cero = [f for f in range(PRIMERA, ULTIMA + 1)
              if r.cell(row=f, column=9).value in (0, None)]
check(not netos_cero, f"ninguna fila con neto en 0 (filas vacías: {netos_cero})")

# --- Referencias fuera de rango --------------------------------------------
print("\nREFERENCIAS")
malas = []
for hoja in (r, sp):
    for fila in hoja.iter_rows():
        for cel in fila:
            v = cel.value
            if isinstance(v, str) and v.startswith("="):
                for ref in re.findall(r"\$?[A-L]\$?(\d+)", v):
                    if int(ref) < 1 or int(ref) > hoja.max_row:
                        malas.append(f"{hoja.title}!{cel.coordinate}: {v}")
check(not malas, f"toda fórmula apunta dentro del rango usado ({len(malas)} fuera)")

prohibidas = ["XLOOKUP","XMATCH","UNIQUE(","FILTER(","SEQUENCE(","TEXTJOIN","_xlfn"]
usadas = [p for p in prohibidas
          for fila in r.iter_rows() for cel in fila
          if isinstance(cel.value, str) and p in cel.value]
check(not usadas, f"sin funciones que LibreOffice no evalúa: {set(usadas) or 'ninguna'}")

# --- Hoja Respaldos ---------------------------------------------------------
print("\nHOJA RESPALDOS")
check(sp["A4"].value == "GASTO N° 1", f"A4 = {sp['A4'].value!r}")
check(sp["A40"].value == "GASTO N° 2", f"A40 (4+36) = {sp['A40'].value!r}")
check(sp["A76"].value == "GASTO N° 3", f"A76 (4+72) = {sp['A76'].value!r}")
check(sp["A5"].value == "Fecha:" and sp["B5"].value == "2026-08-01", "ficha 1: fecha")
check(sp["B14"].value == 50000, f"ficha 1: total = {sp['B14'].value!r}")
aviso = sp["A88"].value
check(isinstance(aviso, str) and "Sin imagen" in aviso,
      f"gasto 3 sin archivo → aviso, no hueco: {str(aviso)[:60]!r}")

print("\nENLACES INTERNOS")
check(r["L12"].hyperlink is not None and "Respaldos" in str(r["L12"].hyperlink.location or r["L12"].hyperlink.target),
      f"L12 → hoja Respaldos ({r['L12'].value})")
check(sp["A15"].hyperlink is not None, "'Volver a la Rendición' es enlace")

print("\nIMÁGENES EMBEBIDAS")
# openpyxl no lee los anclajes que escribe exceljs, así que se inspecciona el
# XML de dibujo directamente: es la fuente autoritativa de dónde queda la imagen.
EMU = 9525  # EMU por píxel
with zipfile.ZipFile(ruta) as z:
    media = [n for n in z.namelist() if n.startswith("xl/media/") and n != "xl/media/"]
    dibujo = z.read("xl/drawings/drawing1.xml").decode()

check(len(media) == 2, f"2 imágenes en el paquete: {media}")

anclajes = re.findall(
    r"<xdr:row>(\d+)</xdr:row>.*?<xdr:ext cx=\"(\d+)\" cy=\"(\d+)\"", dibujo, re.S
)
check(len(anclajes) == 2, f"2 anclajes en el XML de dibujo: {len(anclajes)}")

# base = 4 + i*36, la imagen va en base+12 → filas 16 y 52 (1-indexadas)
for i, (fila0, cx, cy) in enumerate(anclajes):
    fila = int(fila0) + 1
    esperada = 4 + i * 36 + 12
    w, h = round(int(cx) / EMU), round(int(cy) / EMU)
    check(fila == esperada, f"imagen {i+1} anclada en fila {fila} (esperada {esperada})")
    check(0 < w <= 480 and 0 < h <= 520, f"imagen {i+1}: {w}x{h} px dentro de 480x520")

# Proporciones de las boletas de prueba: 1400x900 y 800x1400
esperado_px = [(480, round(480 * 900 / 1400)), (round(520 * 800 / 1400), 520)]
for i, ((_, cx, cy), (we, he)) in enumerate(zip(anclajes, esperado_px)):
    w, h = round(int(cx) / EMU), round(int(cy) / EMU)
    check(abs(w - we) <= 1 and abs(h - he) <= 1,
          f"imagen {i+1}: proporción respetada, {w}x{h} ≈ {we}x{he}")

print("\n" + "=" * 60)
if fallos:
    print(f"{len(fallos)} FALLA(S):")
    for f in fallos:
        print("  -", f)
    sys.exit(1)
print("TODAS LAS COMPROBACIONES PASARON")
